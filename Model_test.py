import shutil
import pandas as pd
import numpy as np
from keras.layers import Dense, Flatten, Conv1D, MaxPooling1D, Input
from keras import backend as K
from openpyxl import Workbook
from keras.models import Model
import os
from sklearn.metrics import r2_score


def test_SO2_model_cnn_single(path, model_stru1, model_para1, test_path, optimizer, loss, num):
    # Loading data
    df = pd.read_pickle(path)
    x_test = np.expand_dims(df.values[:, 0:-2].astype(float), axis=2)  # Adding a one-dimensional axis
    y_test = df.values[:, -2:] / num
    # Divide training set, test set
    print("Loading of data complete!")
    x_test = x_test
    y_test = y_test
    print("y_test: ", y_test)
    loaded_model = load_first_cnn(model_stru1, model_para1)
    loaded_model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = predict_data(loaded_model, x_test, optimizer, loss)
    mean_error(predicted, y_test, num)
    show_scores(predicted, y_test)
    save_model_excel(predicted, y_test, test_path, num)
    print("Test complete!")

def load_first_cnn(model_stru2, model_para1):
    loaded_model = test_build_CNN_model()
    loaded_model.load_weights(model_para1)
    print("Loading of pre-trained models complete!")
    return loaded_model


def test_build_CNN_model():
    input1 = Input(shape=(403, 1))               #!!!!!!!!!!!!!!!!
    conv_layer1_1 = Conv1D(16, 3, strides=1, activation='relu')(input1)
    conv_layer1_2 = Conv1D(16, 3, strides=1, activation='relu')(conv_layer1_1)
    conv_layer1_3 = Conv1D(16, 3, strides=1, activation='relu')(conv_layer1_2)
    conv_layer1_4 = Conv1D(16, 3, strides=1, activation='relu')(conv_layer1_3)
    max_layer1 = MaxPooling1D(3)(conv_layer1_4)
    conv_layer2_1 = Conv1D(32, 3, strides=1, activation='relu')(max_layer1)
    conv_layer2_2 = Conv1D(32, 3, strides=1, activation='relu')(conv_layer2_1)
    conv_layer2_3 = Conv1D(32, 3, strides=1, activation='relu')(conv_layer2_2)
    conv_layer2_4 = Conv1D(32, 3, strides=1, activation='relu')(conv_layer2_3)
    max_layer2 = MaxPooling1D(3)(conv_layer2_4)
    flatten = Flatten()(max_layer2)
    f1 = Dense(2, activation='linear', name='prediction_one')(flatten)
    model = Model(outputs=f1, inputs=input1)
    model.summary()
    return model

def coeff_determination(y_true, y_pred):
    SS_res = K.sum(K.square(y_true - y_pred))
    SS_tot = K.sum(K.square(y_true - K.mean(y_true)))
    return (1 - SS_res / (SS_tot + K.epsilon()))

def mean_error(predicted, y_test, num):
    print(predicted)
    predicted1 = np.reshape(predicted[0], len(predicted[0]))
    predicted2 = np.reshape(predicted[1], len(predicted[1]))
    y_test_size1 = y_test[0]
    y_test_size2 = y_test[1]
    predicted1 = np.array(predicted1) * num
    predicted2 = np.array(predicted2) * num
    y_test_size1 = np.array(y_test_size1) * num
    y_test_size2 = np.array(y_test_size2) * num
    result1 = np.mean(abs((predicted1 - y_test_size1)))
    result2 = np.mean(abs((predicted2 - y_test_size2)))
    print("MAE:{:.2f}".format(result1))
    print("MAE:{:.2f}".format(result2))
    return result1, result2

# Predicted data
def predict_data(model, X_test, optimizer, loss):
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = model.predict(X_test)
    return predicted

def save_model_excel(predicted, Y_test, name, num):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    Y_test = Y_test
    ws.cell(1, 1, 'Predicted concentration')
    ws.cell(1, 2, 'Real concentration')
    for i in range(len(predicted)):
        ws.cell(i + 2, 1, predicted[i][0] * num)
        ws.cell(i + 2, 2, predicted[i][1] * num)
        ws.cell(i + 2, 3, Y_test[i][0] * num)
        ws.cell(i + 2, 4, Y_test[i][1] * num)
    wb.save(name)
    print("Save value to finish!")



def show_scores(predicted, y_test):
    predicted = np.array(predicted)
    print(predicted[0])
    print(y_test[0])
    r2_scores1 = r2_score(predicted[0], y_test[0])

    r2_scores2 = r2_score(predicted[1], y_test[1])
    print("R2_1:", r2_scores1)
    print("R2_2:", r2_scores2)
    return r2_scores1,r2_scores2



def del_files(path):
    if os.path.exists(path):
        shutil.rmtree(path, ignore_errors=False, onerror=None)
    print("Data environment cleanup succeeded!")



if __name__ == '__main__':

    optimizer = "adam"
    loss = "mean_squared_error"
    Test_so2_result_data_path = "Test_SO2_results_data-11"

    del_files(Test_so2_result_data_path)
    os.mkdir(Test_so2_result_data_path)
    Pkl_data_path = "demo/6.pkl"
    result_values_file_name = Test_so2_result_data_path + "/so2-results-11.xlsx"
    model_para = "Result_so2_cnn-11/so2-para.h5"
    model_stru = "Result_so2_cnn-11/so2-stru.json"
    test_SO2_model_cnn_single(Pkl_data_path, model_stru, model_para, result_values_file_name, optimizer, loss, 100)