import os
import random
import numpy as np
import xlwt
import re
import pandas as pd
from openpyxl import Workbook
import shutil
from pandas import DataFrame
# from Sin_reconstruction_method import Sin_reconstruction
import Expansion_one_section_fitting2

def read_single_component_data(path, newpath, txt_num):
    os.mkdir(newpath)
    path = path
    file_name_lists = []
    for file_name in os.listdir(path):
        file_name_lists.append(file_name)
    for file_name_list in file_name_lists:
        root_ = []
        dirs_ = []
        a = 0
        for root, dirs, files in os.walk(path + '\\' + file_name_list):
            root_.append(root)
            dirs_.append(dirs)
            dirs__ = dirs_[0]
        file_name_after = newpath + '\\' + file_name_list + '_average'
        if not os.path.exists(file_name_after):
            os.mkdir(file_name_after)
        # print(root_)
        for i in root_[1:]:
            file_name_after_ = file_name_after + '\\' + dirs__[a]
            if not os.path.exists(file_name_after_):
                os.mkdir(file_name_after_)
            txt_ = []
            for file_name_ in os.listdir(i):
                txt_.append(i + '\\' + file_name_)
            txt_num = txt_num  # 取多少文件
            txt_ = txt_[-txt_num:]
            num_all = []
            for txt_name in txt_:
                contents = []
                with open(txt_name, 'r') as f:
                    for line in f.readlines():
                        line = line.split('\n')
                        line = line[0].split('\t')
                        line = list(map(float, line))
                        contents.append(line)
                for content in contents:
                    num_all.append("%.4f" % (float(content[1]) / txt_num))
                if len(num_all) > len(contents):
                    for ii in range(len(num_all)):
                        if ii < len(contents):
                            num_all[ii] = "%.4f" % (num_all[ii] + float(num_all[ii + len(contents)]))
                        else:
                            num_all.pop()
                num_all = list(map(float, num_all))
                f.close()
                txt_name_after = newpath + '\\' + file_name_list + '_average\\' + dirs__[a] + "\\" + dirs__[a] + ".txt"
                with open(txt_name_after, "w") as ff:
                    for li in num_all:
                        ff.write(str(li) + "\n")
                ff.close()
            a += 1
        print(file_name_list, "Data reading completed！")
    print("All  data read completed！")


def del_files(path1):
    if os.path.exists(path1):
        shutil.rmtree(path1, ignore_errors=False, onerror=None)
    print("Data environment cleanup succeeded!")
import re

def extract_number(s):
    numbers = re.findall(r'\d+', s)
    return int(numbers[0]) if numbers else None


def remove_bd(newpath):
    path = newpath
    nn_ = []
    file_name_lists = []
    root_1 = []
    dirs_1 = []
    root_2 = []
    dirs_2 = []
    data = []
    for file_name in os.listdir(path):
        file_name_lists.append(file_name)

    if file_name_lists:
        first_file_name = file_name_lists[0]
        n_p = path + '\\' + first_file_name + '_removebd'
        os.mkdir(n_p)
        nn_ = [n_p]
    else:
        print("The list is empty.")
        return nn_
    for root, dirs, files in os.walk(newpath + "/" + file_name_lists[0]):
        root_1.append(root)
        dirs_1.append(dirs)

    dirs__1 = dirs_1[0]
    root_1.pop(0)

    root_1_sorted = sorted(root_1, key=lambda x: int(re.findall(r'\d+', os.path.basename(x))[0]))

    for root, dirs, files in os.walk(newpath + "/" + file_name_lists[1]):
        root_2.append(root)
        dirs_2.append(dirs)

    dirs__2 = dirs_2[0]
    root_2.pop(0)

    root_2_sorted = sorted(root_2, key=lambda x: int(re.findall(r'\d+', os.path.basename(x))[0]))
    dirs__1_sorted = sorted(dirs__1, key=extract_number)
    dirs__2_sorted = sorted(dirs__2, key=extract_number)

    for i2 in range(len(root_2)):
        data2 = []
        bd_name = root_2_sorted[i2] + '/' + dirs__2_sorted[i2] + '.txt'
        with open(bd_name, 'r+') as bd_f:
            bd_file = bd_f.read().split('\n')
        bd_f.close()
        del (bd_file[-1])
        bd_file = list(map(float, bd_file))



    for i1 in range(len(root_1)):
        file_name = root_1_sorted[i1] + '/' + dirs__1_sorted[i1] + '.txt'
        file_name_ = n_p + '/' + dirs__1_sorted[i1] + '.txt'
        with open(file_name, 'r+') as f:
            file = f.read().split('\n')
        f.close()
        del (file[-1])
        file = list(map(float, file))

        bd_file = list(map(float, bd_file))
        data = []

        for i in range(len(bd_file)):
            i_num = "%.4f" % ((file[i]) / (bd_file[i]))
            data.append(i_num)
        with open(file_name_, 'w') as f_:
            f_.truncate(0)
            for ii in data:
                f_.write(ii + "\n")
        f_.close()

    print('Successfully removed the backing!')
    return nn_


def writeinexcel(path, nn):
    print(":path:",path)
    lu = []
    path = path
    # print(path)
    le_ = 0
    le = 0
    wb1 = xlwt.Workbook(encoding='utf-8')
    w1 = wb1.add_sheet('one')
    ipath_ = path[0]
    # print(ipath_)
    file_name_lists_ = []
    file_name_lists = []



    for file_name_ in os.listdir(ipath_):
        file_name_ = re.sub('\D', '', file_name_)
        b = list(file_name_)
        # b.pop()
        ans = "".join(map(str, b))
        file_name_lists_.append(ans)
    file_name_lists_ = list(map(int, file_name_lists_))
    print("file_name_lists_:",file_name_lists_)
    file_name_lists_.sort()
    print("file_name_lists_sort:", file_name_lists_)
    le = len(file_name_lists_)
    for le_i in range(le):
        w1.write(0, le_i + le_, file_name_lists_[le_i])
    le_ = le_ + len(file_name_lists_)
    for file_name_ in os.listdir(ipath_):
        file_name_lists.append(file_name_)
    file_name_lists.sort(key=lambda x: int(x[:-nn+3].strip() or '0'))
    for i_a in range(len(file_name_lists)):
        path_ = ipath_ + '\\' + file_name_lists[i_a]
        ii = 1
        for line in open(path_, encoding='utf-8'):
            if line == '\n':
                continue
            else:
                w1.write(ii, i_a, float(line))
                ii += 1

    wb1.save(path[0] + ".xls")
    p1 = path[0] + ".xls"
    lu.append(p1)

    if len(path) > 1:
        ipath__ = path[1]
        le__ = 0
        lei = 0
        wb2 = xlwt.Workbook(encoding='utf-8')
        w2 = wb2.add_sheet('one')
        file_name_lists_ = []
        file_name_lists = []
        for file_name_ in os.listdir(ipath__):
            file_name_ = re.sub('\D', '', file_name_)
            b = list(file_name_)
            ans = "".join(map(str, b))
            file_name_lists_.append(ans)
        file_name_lists_ = list(map(int, file_name_lists_))
        file_name_lists_.sort()
        lei = len(file_name_lists_)
        for le_i in range(lei):
            w2.write(0, le_i + le__, file_name_lists_[le_i])
        le__ = le__ + len(file_name_lists_)
        for file_name_ in os.listdir(ipath__):
            file_name_lists.append(file_name_)
        file_name_lists.sort(key=lambda x: int(x[:-nn].strip() or '0'))
        for i_a in range(len(file_name_lists)):
            path_ = ipath__ + '\\' + file_name_lists[i_a]
            ii = 1
            for line in open(path_, encoding='utf-8'):
                if line == '\n':
                    continue
                else:
                    w2.write(ii, i_a, float(line))
                    ii += 1
        wb2.save(path[1] + ".xls")
        p2 = path[1] + ".xls"
        lu.append(p2)

    return lu


def koumanbian(x):
    aa = []
    for i in x:
        zz1 = np.polyfit([i for i in range(len(i))], i, 3)  # 拟合
        pp1 = np.poly1d(zz1)
        aa.append(np.log(i / (pp1([i for i in range(len(i))]))))
    return np.array(aa)


def dif_SO2(path):
    a = 357
    b = 759

    data = pd.read_excel(path)
    columns = data.columns
    wb_all = Workbook()
    wb_all.create_sheet(index=0, title="all")
    ws_all = wb_all.active

    data_line_1 = data.iloc[a - 1:b, :]

    data_line_1 = np.array(data_line_1)

    data_all = data_line_1
    data_all = pd.DataFrame(data_all, columns=columns)
    data_all = data_all.T
    data_all = np.array(data_all)

    data_deal_all_line = koumanbian(data_all)
    deal_data_line_all = pd.DataFrame(data_deal_all_line)
    columns_ = deal_data_line_all.columns
    l = len(columns_)
    i1 = 0
    for i in range(l):
        i1 += 1
        lie = deal_data_line_all[columns_[i]]
        for j in range(len(lie)):
            ws_all.cell(1, i + 1, i1)
            ws_all.cell(j + 2, i + 1, lie[j])
    for col_i in range(len(columns)):
        col = columns[col_i]
        ws_all.cell(col_i + 2, l + 1, col)
    ws_all.cell(1, l + 1, l + 1)
    path = path.split('.')[0]
    name = path + '.xlsx'
    wb_all.save(name)
    print("SO2 Data differential completion!")
    return name



def to_pkl(path):
    df1 = DataFrame(pd.read_excel(path))
    dir_name = os.path.dirname(path)
    base_name = os.path.basename(path)
    suffix = base_name.split(".")[0]
    path_ = dir_name + "/" + suffix + ".pkl"
    df1.to_pickle(path_)
    return path_


def UV_DOAS():
    Raw_data_path = "Raw_SO2_data129"
    Processed_data_path = "Processed_SO2_data129"
    so2_spectrum_path = Processed_data_path + '550-SO2_average_removebd.xlsx'
    txt_num = 50
    numb = 9
    # del_files(Processed_data_path)
    # read_single_component_data(Raw_data_path, Processed_data_path, txt_num)
    # print("Processed_data_path:",Processed_data_path)
    # path1 = remove_bd(Processed_data_path)
    # path2 = writeinexcel(path1, numb)
    # path3 = dif_SO2(path2[0])
    # print(path3)
    path3 = "Processed_SO2_data129/550-SO2_average_removebd.xlsx"
    Expansion_one_section_fitting2.run(path3)




if __name__ == '__main__':
    UV_DOAS()
