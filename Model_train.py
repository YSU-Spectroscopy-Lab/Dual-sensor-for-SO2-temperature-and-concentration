import Model_build
import os
import shutil
from keras import backend as K
from openpyxl import Workbook
import pandas as pd
import numpy as np
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt

def train_model(model, name, loss_name, model_para, model_stru, optimizer, loss, X_train, X_test, Y_train, Y_test,
                name1, name2, MAE_r2_name):
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])

    history = model.fit(X_train, Y_train,
                        batch_size=256,
                        epochs=500, #1000  /  1   / 500  /100
                        validation_data=(X_test, Y_test)
                        )
    predicted = predict_data(model, X_test, optimizer=optimizer, loss=loss)
    MAE1,MAE2 = mean_error(predicted, Y_test, num)
    save_model(model, model_stru, model_para)
    r2_1,r2_2 = show_scores(predicted, Y_test)
    save_model_excel(predicted, Y_test, name, num)
    save_MAE_r2(MAE1,MAE2, r2_1,r2_2, MAE_r2_name)
    return MAE1,MAE2

def coeff_determination(y_true, y_pred):
    SS_res = K.sum(K.square(y_true - y_pred))
    SS_tot = K.sum(K.square(y_true - K.mean(y_true)))
    return (1 - SS_res / (SS_tot + K.epsilon()))


def save_model_excel(predicted, Y_test, name, num):
    wb = Workbook()
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    Y_test = Y_test
    ws.cell(1, 1, 'Predicted concentration')
    ws.cell(1, 2, 'Real concentration')
    for i in range(len(predicted)):
        ws.cell(i + 2, 1, predicted[i][0] * num)
        ws.cell(i + 2, 2, predicted[i][1] * num)
        ws.cell(i + 2, 3, Y_test[i][0] * num)
        ws.cell(i + 2, 4, Y_test[i][1] * num)
    wb.save(name)
    print("Save value to finish!")


def save_model_loss(history, name):
    wb = Workbook()
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    hist = pd.DataFrame(history.history)
    loss = hist['loss']
    val_loss = hist['val_loss']
    ws.cell(1, 1, 'loss')
    ws.cell(1, 2, 'val_loss')
    for i in range(len(loss)):
        ws.cell(i + 2, 1, loss[i])
        ws.cell(i + 2, 2, val_loss[i])
    wb.save(name)
    print("Save loss to finish!")

def predict_data(model, X_test, optimizer, loss):
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = model.predict(X_test)
    return predicted

def show_scores(predicted, Y_test):
    predicted = np.array(predicted)
    r2_scores1 = r2_score(predicted[0], Y_test[0])
    r2_scores2 = r2_score(predicted[1], Y_test[1])
    print("R2_1:", r2_scores1)
    print("R2_2:", r2_scores2)
    return r2_scores1,r2_scores2

def mean_error(predicted, y_test, num):
    print(predicted)
    predicted1 = np.reshape(predicted[0], len(predicted[0]))
    predicted2 = np.reshape(predicted[1], len(predicted[1]))
    y_test_size1 = y_test[0]
    y_test_size2 = y_test[1]

    predicted1 = np.array(predicted1) * num
    predicted2 = np.array(predicted2) * num
    y_test_size1 = np.array(y_test_size1) * num

    y_test_size2 = np.array(y_test_size2) * num

    result1 = np.mean(abs((predicted1 - y_test_size1)))
    result2 = np.mean(abs((predicted2 - y_test_size2)))
    print("MAE:{:.2f}".format(result1))
    print("MAE:{:.2f}".format(result2))
    return result1, result2


# Preservation of models
def save_model(model, name1, name2):
    model_json = model.to_json()
    with open(name1, 'w') as json_file:
        json_file.write(model_json)
    model.save_weights(name2)
    print("Save model complete!")


def plot_history(history, name1, name2):
    hist = pd.DataFrame(history.history)
    hist['epoch'] = history.epoch
    plt.figure()
    plt.xlabel('Epoch')
    plt.ylabel('loss')
    plt.plot(hist['epoch'], hist['loss'],
             label='Train loss')
    plt.plot(hist['epoch'], hist['val_loss'],
             label='Val loss')
    plt.ylim([-0.001, 0.05])
    plt.legend()
    plt.savefig(name1, dpi=600)
    plt.figure()
    plt.xlabel('Epoch')
    plt.ylabel('coeff_determination')
    plt.plot(hist['epoch'], hist['coeff_determination'],
             label='Train coeff_determination')
    plt.plot(hist['epoch'], hist['val_coeff_determination'],
             label='Val coeff_determination')
    plt.ylim([-0.5, 1.5])
    plt.legend()
    plt.savefig(name2, dpi=600)
    # plt.show()


def save_MAE_r2(MAE1,MAE2, r2_1,r2_2, name):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active

    ws.cell(1, 1, 'MAE1')
    ws.cell(1, 2, 'MAE2')
    ws.cell(1, 3, 'R1')
    ws.cell(1, 4, 'R2')
    ws.cell(2, 1, MAE1)
    ws.cell(2, 2, MAE2)
    ws.cell(2, 3, r2_1)
    ws.cell(2, 4, r2_2)
    wb.save(name)
    print("Save MAE & r2 to finish!")


def del_files(path):
    if os.path.exists(path):
        shutil.rmtree(path, ignore_errors=False, onerror=None)
    print("Data environment cleanup succeeded!")



if __name__ == '__main__':
    Result_so2_data_path = "Result_so2_cnn"

    optimizer = "adam"
    loss = "mean_squared_error"

    model_structure = Result_so2_data_path + "/CNN.png"
    result_values_file_name = Result_so2_data_path + "/so2-results.xlsx"
    loss_file_name = Result_so2_data_path + "/so2-loss.xlsx"
    model_para = Result_so2_data_path + "/so2-para.h5"
    model_stru = Result_so2_data_path + "/so2-stru.json"
    model_loss = Result_so2_data_path + "/so2-loss.png"
    model_R2 = Result_so2_data_path + "/so2-R2.png"
    model_MAE_r2_name = Result_so2_data_path + "/so2-MAE-r2.xlsx"


    del_files(Result_so2_data_path)
    os.mkdir(Result_so2_data_path)
    model, X_train, X_test, Y_train, Y_test, num = Model_build.run(model_structure)
    MAE = train_model(model, result_values_file_name, loss_file_name, model_para, model_stru, optimizer, loss, X_train,
                X_test, Y_train,
                Y_test, model_loss, model_R2, model_MAE_r2_name)

